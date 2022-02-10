import csv
from openpyxl import load_workbook
from dao.clae import ClaeDAO
from dao.empresa import EmpresaDAO
from dao.planta import PlantaDAO

def cargar_clae():
    clae = "clae.txt"
    with open(clae, encoding="utf8") as archivo:
        lector = csv.reader(archivo, delimiter=";")
        next(lector, None)
        for fila in lector:
            dic = {
                'id_clae' : fila[0],
                'actividad' : fila[1]
            }
            ClaeDAO.insertar(dic)

def prueba():
    libro = load_workbook('datos.xlsx')
    # hoja = libro['empresas']
    # fila = hoja.max_row 
    # columna = hoja.max_column


    # print('EMPRESAS')
    # for x in range(2, fila+1):
    #     lista = []
    #     for y in range(2, columna+1):
    #         if y != 10 and y != 11:
    #             celda = hoja.cell(x, y).value
    #             if type(celda) != str:
    #                 celda = str(celda)
    #             lista.append(celda)
    #     EmpresaDAO.insertar(lista)


    hoja = libro['plantas']
    fila = hoja.max_row 
    columna = hoja.max_column


    print('\nPLANTAS\n')
    for x in range(2, fila+1):
        lista = []
        for y in range(2, columna+1):
            celda = hoja.cell(x, y).value
            
            if (y in [4, 5, 9]) and (type(celda) != str and celda != None):
                celda = str(celda)
            lista.append(celda)
        PlantaDAO.insertar(lista)            

prueba()